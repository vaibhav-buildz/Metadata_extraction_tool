from .base import MetadataExtractor
from typing import Dict, Any
from docx import Document
from openpyxl import load_workbook


class OfficeMetadataExtractor(MetadataExtractor):
    supported_extensions = ('.docx', '.xlsx', '.pptx')
    
    def extract(self) -> Dict[str, Any]:
        if self.file_path.suffix.lower() == '.docx':
            return self._extract_docx()
        elif self.file_path.suffix.lower() == '.xlsx':
            return self._extract_xlsx()
        elif self.file_path.suffix.lower() == '.pptx':
            return self._extract_pptx()
        else:
            return {
                "file_type": "Unknown Office Document",
                "filename": self.filename,
                "file_size_bytes": self.file_size,
                "status": "Unsupported office document format",
                "anomalies": []
            }
    
    def _parse_ooxml_date(self, date_str: str):
        """Parse standard ISO/OOXML datetime strings without external dependencies"""
        if not date_str or date_str == "Not set":
            return None
        # Clean suffix and extract base
        cleaned = date_str.replace('Z', '').split('.')[0]
        from datetime import datetime
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                return datetime.strptime(cleaned, fmt)
            except ValueError:
                continue
        return None

    def _check_office_anomalies(self, metadata: Dict[str, Any]):
        """Perform forensic audit on OOXML document metadata"""
        # 1. Check for macros (VBA macro project file)
        try:
            import zipfile
            if zipfile.is_zipfile(self.file_path):
                with zipfile.ZipFile(self.file_path, 'r') as z:
                    names = z.namelist()
                    if any("vbaProject.bin" in name for name in names):
                        metadata["anomalies"].append("Embedded VBA Macros detected (vbaProject.bin is present, potential active malware payload)")
        except Exception:
            pass
            
        # 2. Created vs modified date comparison
        core = metadata.get("core_properties") or metadata.get("workbook_properties")
        if core:
            created = core.get("created")
            modified = core.get("modified")
            c_dt = self._parse_ooxml_date(created)
            m_dt = self._parse_ooxml_date(modified)
            if c_dt and m_dt and m_dt < c_dt:
                metadata["anomalies"].append(f"Creation timestamp tampering detected: File modification date ({modified}) is BEFORE creation date ({created})")
            
            # 3. Creator vs last modified author mismatch
            author = core.get("author") or core.get("creator")
            last_author = core.get("last_modified_by")
            if author and last_author and author != "Not set" and last_author != "Not set":
                if author.lower() != last_author.lower():
                    metadata["anomalies"].append(f"Collaborator/Author mismatch: Document created by '{author}' but last modified by '{last_author}'")

    def _extract_docx(self) -> Dict[str, Any]:
        metadata = {
            "file_type": "Word Document",
            "filename": self.filename,
            "file_size_bytes": self.file_size,
            "core_properties": {},
            "anomalies": []
        }
        
        try:
            doc = Document(str(self.file_path))
            props = doc.core_properties
            metadata["core_properties"] = {
                "author": props.author or "Not set",
                "last_modified_by": props.last_modified_by or "Not set",
                "title": props.title or "Not set",
                "subject": props.subject or "Not set",
                "keywords": props.keywords or "Not set",
                "revision": str(props.revision) if props.revision is not None else "Not set",
                "created": str(props.created) if props.created else "Not set",
                "modified": str(props.modified) if props.modified else "Not set",
            }
        except Exception as e:
            metadata["core_properties"]["error"] = str(e)
            metadata["anomalies"].append(f"DOCX property parsing failed: {e}")
        
        self._check_office_anomalies(metadata)
        return metadata
    
    def _extract_xlsx(self) -> Dict[str, Any]:
        metadata = {
            "file_type": "Excel Spreadsheet",
            "filename": self.filename,
            "file_size_bytes": self.file_size,
            "workbook_properties": {},
            "sheets": [],
            "sheets_summary": {},
            "anomalies": []
        }
        
        try:
            wb = load_workbook(str(self.file_path))
            props = wb.properties
            metadata["workbook_properties"] = {
                "creator": props.creator or "Not set",
                "last_modified_by": props.lastModifiedBy or "Not set",
                "title": props.title or "Not set",
                "created": str(props.created) if props.created else "Not set",
                "modified": str(props.modified) if props.modified else "Not set",
            }
            
            # Sheets analysis with hidden sheet audit
            hidden_sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                state = getattr(ws, 'sheet_state', 'visible')
                if state in ('hidden', 'veryHidden'):
                    hidden_sheets.append(f"{sheet_name} ({state})")
                
                metadata["sheets"].append({
                    "name": sheet_name,
                    "rows": ws.max_row,
                    "columns": ws.max_column,
                    "state": state
                })
                metadata["sheets_summary"][sheet_name] = f"{ws.max_row} rows x {ws.max_column} cols ({state})"
                
            if hidden_sheets:
                metadata["anomalies"].append(f"Hidden spreadsheet tabs detected: {', '.join(hidden_sheets)} (potential hidden data repository)")
        except Exception as e:
            metadata["workbook_properties"]["error"] = str(e)
            metadata["anomalies"].append(f"XLSX sheet/property parsing failed: {e}")
        
        self._check_office_anomalies(metadata)
        return metadata

    def _extract_pptx(self) -> Dict[str, Any]:
        metadata = {
            "file_type": "PowerPoint Presentation",
            "filename": self.filename,
            "file_size_bytes": self.file_size,
            "core_properties": {},
            "presentation_properties": {},
            "anomalies": []
        }
        
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            
            with zipfile.ZipFile(self.file_path, 'r') as z:
                # 1. Parse core properties (docProps/core.xml)
                try:
                    core_xml = z.read('docProps/core.xml')
                    root = ET.fromstring(core_xml)
                    
                    ns = {
                        'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties',
                        'dc': 'http://purl.org/dc/elements/1.1/',
                        'dcterms': 'http://purl.org/dc/terms/',
                        'xsi': 'http://www.w3.org/2001/XMLSchema-instance'
                    }
                    
                    def find_val(xpath):
                        el = root.find(xpath, ns)
                        return el.text if el is not None else None
                    
                    metadata["core_properties"] = {
                        "author": find_val('dc:creator') or "Not set",
                        "last_modified_by": find_val('cp:lastModifiedBy') or "Not set",
                        "title": find_val('dc:title') or "Not set",
                        "created": find_val('dcterms:created') or "Not set",
                        "modified": find_val('dcterms:modified') or "Not set",
                        "revision": find_val('cp:revision') or "Not set",
                    }
                except Exception as e:
                    metadata["core_properties"]["error"] = f"Failed to parse docProps/core.xml: {e}"
                
                # 2. Parse app properties (docProps/app.xml)
                try:
                    app_xml = z.read('docProps/app.xml')
                    root = ET.fromstring(app_xml)
                    
                    ns_app = {'ep': 'http://schemas.openxmlformats.org/officeDocument/2006/extended-properties'}
                    
                    slides_el = root.find('ep:Slides', ns_app)
                    slides = int(slides_el.text) if slides_el is not None else 0
                    
                    notes_el = root.find('ep:Notes', ns_app)
                    notes = int(notes_el.text) if notes_el is not None else 0
                    
                    presentation = root.find('ep:PresentationFormat', ns_app)
                    fmt = presentation.text if presentation is not None else "Unknown"
                    
                    metadata["presentation_properties"] = {
                        "slide_count": slides,
                        "notes_count": notes,
                        "presentation_format": fmt
                    }
                except Exception as e:
                    metadata["presentation_properties"]["error"] = f"Failed to parse docProps/app.xml: {e}"
        except Exception as e:
            metadata["anomalies"].append(f"PPTX Zip parsing failed: {e}")
            
        self._check_office_anomalies(metadata)
        return metadata


if __name__ == "__main__":
    print("✓ OfficeMetadataExtractor ready")

